#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date

def create_edge_case_excel():
    """예외 케이스가 포함된 엑셀 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "직원 정보"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 작성
    headers = ["이름", "나이", "부서", "직급", "연봉", "입사일", "최근 로그인", "재직여부"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # 예외 케이스 데이터
    data = [
        # 1. 정상 데이터
        ["김철수", 32, "개발팀", "대리", 5500, date(2020, 3, 15), datetime(2025, 12, 8, 9, 30, 0), True],

        # 2. 나이가 문자열로 되어있는 경우
        ["이영희", "28세", "마케팅팀", "사원", 4200, date(2022, 6, 1), datetime(2025, 12, 8, 8, 45, 0), True],

        # 3. 연봉이 문자열로 되어있는 경우 (쉼표 포함)
        ["박민수", 35, "개발팀", "과장", "7,000원", date(2018, 1, 10), datetime(2025, 12, 7, 18, 20, 0), True],

        # 4. 날짜가 문자열로 되어있는 경우
        ["정수진", 30, "인사팀", "대리", 5200, "2021-09-20", "2025-12-08 10:15:00", True],

        # 5. Boolean이 문자열로 되어있는 경우
        ["최동욱", 42, "영업팀", "부장", 9000, date(2015, 4, 5), datetime(2025, 12, 8, 7, 30, 0), "TRUE"],

        # 6. 빈 셀이 섞여있는 경우
        ["강민지", None, "디자인팀", "사원", 4000, date(2023, 3, 1), None, True],

        # 7. 나이가 실수로 되어있는 경우
        ["윤서현", 33.5, "개발팀", "과장", 6800, date(2019, 7, 15), datetime(2025, 12, 8, 8, 0, 0), True],

        # 8. 모든 필드가 문자열인 경우
        ["한지훈", "29", "마케팅팀", "대리", "5300", "2021/11/10", "2025/12/08 09:45:00", "Yes"],

        # 9. 날짜 형식이 다른 경우 (dd/MM/yyyy)
        ["오현아", 31, "영업팀", "대리", 5600, "20/05/2020", datetime(2025, 12, 7, 17, 30, 0), True],

        # 10. 특수문자가 포함된 경우
        ["임태양", "38세 (만)", "개발팀", "차장", 8200.5, date(2016, 8, 1), datetime(2025, 12, 8, 8, 30, 0), True],

        # 11. 날짜가 잘못된 형식인 경우
        ["조민재", 27, "기획팀", "사원", 4500, "invalid-date", datetime(2024, 1, 15, 9, 0, 0), False],

        # 12. 숫자에 공백이 포함된 경우
        ["서지우", "  35  ", "개발팀", "과장", "  6500  ", date(2019, 3, 10), datetime(2025, 12, 8, 9, 15, 0), True],
    ]

    # 데이터 작성 (타입을 그대로 유지)
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)

            # 값을 문자열로 저장할 케이스
            if row_num == 3 and col_num == 2:  # "28세"
                cell.value = value
            elif row_num == 4 and col_num == 5:  # "7,000원"
                cell.value = value
            elif row_num == 5 and col_num in [6, 7]:  # 날짜 문자열
                cell.value = value
            elif row_num == 6 and col_num == 8:  # "TRUE"
                cell.value = value
            elif row_num == 9 and col_num in [2, 3, 4, 5, 6, 7, 8]:  # 모든 필드 문자열
                cell.value = str(value) if value is not None else None
            elif row_num == 10 and col_num == 6:  # "20/05/2020"
                cell.value = value
            elif row_num == 11 and col_num == 2:  # "38세 (만)"
                cell.value = value
            elif row_num == 12 and col_num == 6:  # "invalid-date"
                cell.value = value
            elif row_num == 13 and col_num in [2, 5]:  # 공백 포함 숫자
                cell.value = value
            else:
                cell.value = value

                # 날짜 및 시간 포맷 적용 (정상 케이스만)
                if isinstance(value, date) and not isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD'
                elif isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'

    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 파일 저장
    wb.save("sample/직원정보_예외케이스.xlsx")
    print("예외 케이스 엑셀 파일이 생성되었습니다: sample/직원정보_예외케이스.xlsx")
    print("\n포함된 예외 케이스:")
    print("  1. 정상 데이터 (기준)")
    print("  2. 나이가 문자열 ('28세')")
    print("  3. 연봉이 문자열 ('7,000원')")
    print("  4. 날짜가 문자열 ('2021-09-20')")
    print("  5. Boolean이 문자열 ('TRUE')")
    print("  6. 빈 셀 (null)")
    print("  7. 나이가 실수 (33.5)")
    print("  8. 모든 필드가 문자열")
    print("  9. 날짜 형식이 다름 ('20/05/2020')")
    print(" 10. 특수문자 포함 ('38세 (만)')")
    print(" 11. 잘못된 날짜 형식 ('invalid-date')")
    print(" 12. 공백 포함 숫자 ('  35  ', '  6500  ')")

if __name__ == "__main__":
    create_edge_case_excel()
