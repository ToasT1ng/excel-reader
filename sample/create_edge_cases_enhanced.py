#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date

def create_enhanced_edge_case_excel():
    """null 및 다양한 예외 케이스가 포함된 엑셀 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "직원 정보"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 작성
    headers = ["이름", "나이", "부서", "직급", "연봉", "입사일", "최근 로그인", "재직여부"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # 예외 케이스 데이터
    data = [
        # 1. 정상 데이터 (기준)
        ["김철수", 32, "개발팀", "대리", 5500, date(2020, 3, 15), datetime(2025, 12, 8, 9, 30, 0), True],

        # 2. 나이만 null
        ["이영희", None, "마케팅팀", "사원", 4200, date(2022, 6, 1), datetime(2025, 12, 8, 8, 45, 0), True],

        # 3. 부서와 직급이 null
        ["박민수", 35, None, None, 7000, date(2018, 1, 10), datetime(2025, 12, 7, 18, 20, 0), True],

        # 4. 연봉만 null
        ["정수진", 30, "인사팀", "대리", None, date(2021, 9, 20), datetime(2025, 12, 8, 10, 15, 0), True],

        # 5. 입사일만 null
        ["최동욱", 42, "영업팀", "부장", 9000, None, datetime(2025, 12, 8, 7, 30, 0), True],

        # 6. 최근 로그인만 null
        ["강민지", 26, "디자인팀", "사원", 4000, date(2023, 3, 1), None, True],

        # 7. 재직여부만 null
        ["윤서현", 33, "개발팀", "과장", 6800, date(2019, 7, 15), datetime(2025, 12, 8, 8, 0, 0), None],

        # 8. 여러 필드가 null (나이, 연봉, 최근 로그인)
        ["한지훈", None, "마케팅팀", "대리", None, date(2021, 11, 10), None, True],

        # 9. 대부분 필드가 null (이름과 부서만 존재)
        ["오현아", None, "영업팀", None, None, None, None, None],

        # 10. 모든 필드가 null인 행
        [None, None, None, None, None, None, None, None],

        # 11. 첫 번째 필드(이름)만 null
        [None, 38, "개발팀", "차장", 8200, date(2016, 8, 1), datetime(2025, 12, 8, 8, 30, 0), True],

        # 12. 마지막 필드(재직여부)만 null
        ["조민재", 27, "기획팀", "사원", 4500, date(2024, 1, 15), datetime(2024, 1, 15, 9, 0, 0), None],

        # 13. 연속된 null 필드 (부서, 직급, 연봉)
        ["서지우", 35, None, None, None, date(2019, 3, 10), datetime(2025, 12, 8, 9, 15, 0), True],

        # 14. 빈 문자열 (null과 구분)
        ["", 29, "", "대리", 5300, date(2022, 2, 20), datetime(2025, 12, 8, 10, 0, 0), True],

        # 15. 나이가 문자열로 되어있는 경우
        ["김미래", "28세", "개발팀", "사원", 4500, date(2023, 5, 10), datetime(2025, 12, 8, 9, 0, 0), True],

        # 16. 연봉이 문자열로 되어있는 경우 (쉼표 포함)
        ["이과거", 40, "영업팀", "과장", "8,500,000원", date(2017, 8, 20), datetime(2025, 12, 8, 8, 30, 0), True],

        # 17. 날짜가 문자열로 되어있는 경우
        ["박현재", 32, "인사팀", "대리", 5600, "2020-06-15", "2025-12-08 09:30:00", True],

        # 18. Boolean이 문자열로 되어있는 경우
        ["정미래", 35, "기획팀", "과장", 7200, date(2018, 11, 5), datetime(2025, 12, 8, 8, 15, 0), "TRUE"],

        # 19. 나이가 실수로 되어있는 경우
        ["최현실", 33.5, "디자인팀", "대리", 5400, date(2021, 4, 10), datetime(2025, 12, 8, 9, 45, 0), True],

        # 20. 모든 필드가 문자열인 경우
        ["한문자", "30", "개발팀", "대리", "5500", "2021/03/01", "2025/12/08 09:00:00", "Yes"],

        # 21. 날짜 형식이 다른 경우 (dd/MM/yyyy)
        ["윤형식", 31, "영업팀", "대리", 5600, "20/05/2020", datetime(2025, 12, 7, 17, 30, 0), True],

        # 22. 특수문자가 포함된 경우
        ["임특수", "38세 (만)", "개발팀", "차장", 8200.5, date(2016, 8, 1), datetime(2025, 12, 8, 8, 30, 0), True],

        # 23. 날짜가 잘못된 형식인 경우
        ["조오류", 27, "기획팀", "사원", 4500, "invalid-date", datetime(2024, 1, 15, 9, 0, 0), False],

        # 24. 숫자에 공백이 포함된 경우
        ["서공백", "  35  ", "개발팀", "과장", "  6500  ", date(2019, 3, 10), datetime(2025, 12, 8, 9, 15, 0), True],

        # 25. Boolean이 0/1로 되어있는 경우
        ["김숫자", 29, "마케팅팀", "사원", 4300, date(2023, 7, 1), datetime(2025, 12, 8, 8, 0, 0), 1],

        # 26. Boolean이 False인 경우 (퇴사자)
        ["이퇴사", 45, "개발팀", "부장", 9500, date(2010, 3, 1), datetime(2024, 6, 30, 18, 0, 0), False],

        # 27. 음수 나이 (오류 데이터)
        ["박음수", -5, "영업팀", "사원", 4000, date(2024, 1, 1), datetime(2025, 12, 8, 9, 0, 0), True],

        # 28. 매우 큰 숫자 (연봉)
        ["정억만", 50, "임원", "사장", 999999999, date(2005, 1, 1), datetime(2025, 12, 8, 7, 0, 0), True],

        # 29. 0 값들
        ["최제로", 0, "인턴팀", "인턴", 0, date(2025, 12, 1), datetime(2025, 12, 8, 10, 0, 0), True],

        # 30. null과 빈 문자열이 섞인 경우
        ["", None, "", None, "", None, None, ""],
    ]

    # 데이터 작성
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value

            # 날짜 및 시간 포맷 적용 (정상 케이스만)
            if isinstance(value, date) and not isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'

    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(12, (max_length + 2) * 1.2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 파일 저장
    filename = "sample/직원정보_null_edge_cases.xlsx"
    wb.save(filename)

    print(f"✅ null 및 예외 케이스 엑셀 파일이 생성되었습니다: {filename}")
    print("\n📋 포함된 Edge Cases:")
    print("\n[NULL 관련 케이스]")
    print("  1. 정상 데이터 (기준)")
    print("  2. 나이만 null")
    print("  3. 부서와 직급이 null")
    print("  4. 연봉만 null")
    print("  5. 입사일만 null")
    print("  6. 최근 로그인만 null")
    print("  7. 재직여부만 null")
    print("  8. 여러 필드가 null (나이, 연봉, 최근 로그인)")
    print("  9. 대부분 필드가 null (이름과 부서만 존재)")
    print(" 10. 모든 필드가 null인 행")
    print(" 11. 첫 번째 필드(이름)만 null")
    print(" 12. 마지막 필드(재직여부)만 null")
    print(" 13. 연속된 null 필드 (부서, 직급, 연봉)")
    print(" 14. 빈 문자열 (null과 구분)")
    print("\n[타입 관련 케이스]")
    print(" 15. 나이가 문자열 ('28세')")
    print(" 16. 연봉이 문자열 ('8,500,000원')")
    print(" 17. 날짜가 문자열 ('2020-06-15')")
    print(" 18. Boolean이 문자열 ('TRUE')")
    print(" 19. 나이가 실수 (33.5)")
    print(" 20. 모든 필드가 문자열")
    print("\n[형식 관련 케이스]")
    print(" 21. 날짜 형식이 다름 ('20/05/2020')")
    print(" 22. 특수문자 포함 ('38세 (만)')")
    print(" 23. 잘못된 날짜 형식 ('invalid-date')")
    print(" 24. 공백 포함 숫자 ('  35  ', '  6500  ')")
    print(" 25. Boolean이 0/1로 되어있는 경우")
    print(" 26. Boolean이 False인 경우 (퇴사자)")
    print("\n[경계값 케이스]")
    print(" 27. 음수 나이 (-5)")
    print(" 28. 매우 큰 숫자 (999999999)")
    print(" 29. 0 값들")
    print(" 30. null과 빈 문자열이 섞인 경우")
    print(f"\n총 {len(data)}개의 테스트 케이스")

if __name__ == "__main__":
    create_enhanced_edge_case_excel()
