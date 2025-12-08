#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date

def create_sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "직원 정보"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 작성
    headers = ["이름", "나이", "부서", "직급", "연봉", "입사일", "최근 로그인", "재직여부"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # 데이터 작성
    data = [
        ["김철수", 32, "개발팀", "대리", 5500, date(2020, 3, 15), datetime(2025, 12, 8, 9, 30, 0), True],
        ["이영희", 28, "마케팅팀", "사원", 4200, date(2022, 6, 1), datetime(2025, 12, 8, 8, 45, 0), True],
        ["박민수", 35, "개발팀", "과장", 7000, date(2018, 1, 10), datetime(2025, 12, 7, 18, 20, 0), True],
        ["정수진", 30, "인사팀", "대리", 5200, date(2021, 9, 20), datetime(2025, 12, 8, 10, 15, 0), True],
        ["최동욱", 42, "영업팀", "부장", 9000, date(2015, 4, 5), datetime(2025, 12, 8, 7, 30, 0), True],
        ["강민지", 26, "디자인팀", "사원", 4000, date(2023, 3, 1), datetime(2025, 12, 8, 9, 0, 0), True],
        ["윤서현", 33, "개발팀", "과장", 6800, date(2019, 7, 15), datetime(2025, 12, 8, 8, 0, 0), True],
        ["한지훈", 29, "마케팅팀", "대리", 5300, date(2021, 11, 10), datetime(2025, 12, 8, 9, 45, 0), True],
        ["오현아", 31, "영업팀", "대리", 5600, date(2020, 5, 20), datetime(2025, 12, 7, 17, 30, 0), True],
        ["임태양", 38, "개발팀", "차장", 8200, date(2016, 8, 1), datetime(2025, 12, 8, 8, 30, 0), True]
    ]

    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            
            # 날짜 및 시간 포맷 적용
            if isinstance(value, date) and not isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'

    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 파일 저장
    wb.save("sample/직원정보.xlsx")
    print("샘플 엑셀 파일이 생성되었습니다: sample/직원정보.xlsx")

if __name__ == "__main__":
    create_sample_excel()
